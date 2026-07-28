import base64
import io
import zipfile

import pytest
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook, load_workbook

from reports.excel_contacts.application import MAX_FILE_SIZE
from reports.excel_contacts.domain import normalize_phone, process_contact_rows
from reports.models import (
    AuditEvent,
    Service,
    ServiceCategory,
    UserCategoryRestriction,
    UserServiceRestriction,
)

pytestmark = pytest.mark.django_db
URL = "/api/tools/excel-contacts/process/"


@pytest.fixture
def excel_service():
    category = ServiceCategory.objects.create(name="Productivity", slug="productivity")
    return Service.objects.create(
        category=category,
        name="Excel Contacts",
        slug="whatsapp-contacts",
        description="Prepare contacts",
        kind=Service.Kind.INTERNAL,
        launch_target="/tools/excel-contacts",
    )


def make_xlsx(rows=None, *, excessive_row=False):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "جهات الاتصال"
    sheet.append(["الاسم الكامل", "رقم التواصل", "البريد الالكتروني"])
    for row in rows or [["مستخدم تجريبي", "0999123456", "user@example.com"]]:
        sheet.append(row)
    if excessive_row:
        sheet.cell(row=10_002, column=1, value="too many")
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def upload(data=None, name="contacts.xlsx"):
    return SimpleUploadedFile(
        name,
        data if data is not None else make_xlsx(),
        content_type="application/octet-stream",
    )


def authenticate(api, user):
    api.force_authenticate(user=user)


def test_previous_phone_normalization_behavior_is_preserved():
    for value in [
        "0933123456",
        "933123456",
        "00963933123456",
        "+963 933 123 456",
        "٠٩٣٣١٢٣٤٥٦",
    ]:
        assert normalize_phone(value, "963") == "+963933123456"


def test_previous_duplicate_and_invalid_classification_is_preserved():
    result = process_contact_rows(
        [
            {
                "الاسم الكامل": "أحمد الأول",
                "رقم التواصل": "0933123456",
                "البريد الالكتروني": "first@example.com",
            },
            {
                "الاسم الكامل": "أحمد المكرر",
                "رقم التواصل": "+963933123456",
                "البريد الالكتروني": "second@example.com",
            },
            {
                "الاسم الكامل": "ليلى",
                "رقم التواصل": "0944123456",
                "البريد الالكتروني": "invalid-email",
            },
        ]
    )
    assert result["summary"] == {
        "totalRows": 3,
        "validCount": 1,
        "duplicateCount": 1,
        "invalidCount": 1,
    }
    assert result["valid_rows"][0]["الاسم الكامل"] == "أحمد الأول"


def test_unauthenticated_processing_is_rejected(api, excel_service):
    response = api.post(URL, {"file": upload(), "countryCode": "963"}, format="multipart")
    assert response.status_code == 401


def test_inactive_user_is_rejected(api, normal_user, excel_service):
    normal_user.is_active = False
    normal_user.save(update_fields=["is_active"])
    authenticate(api, normal_user)
    response = api.post(URL, {"file": upload()}, format="multipart")
    assert response.status_code == 403
    assert response.data["code"] == "SERVICE_ACCESS_DENIED"


def test_disabled_service_is_rejected(api, normal_user, excel_service):
    excel_service.is_active = False
    excel_service.save(update_fields=["is_active"])
    authenticate(api, normal_user)
    response = api.post(URL, {"file": upload()}, format="multipart")
    assert response.status_code == 403


def test_direct_service_restriction_is_enforced(api, normal_user, excel_service):
    UserServiceRestriction.objects.create(user=normal_user, service=excel_service)
    authenticate(api, normal_user)
    response = api.post(URL, {"file": upload()}, format="multipart")
    assert response.status_code == 403


def test_category_restriction_is_enforced(api, normal_user, excel_service):
    UserCategoryRestriction.objects.create(user=normal_user, category=excel_service.category)
    authenticate(api, normal_user)
    response = api.post(URL, {"file": upload()}, format="multipart")
    assert response.status_code == 403


def test_valid_authenticated_request_succeeds_and_is_audited(api, normal_user, excel_service):
    authenticate(api, normal_user)
    response = api.post(URL, {"file": upload(), "countryCode": "963"}, format="multipart")
    assert response.status_code == 200
    assert response["Cache-Control"] == "no-store"
    assert response.data["summary"] == {
        "totalRows": 1,
        "validCount": 1,
        "duplicateCount": 0,
        "invalidCount": 0,
    }
    with zipfile.ZipFile(io.BytesIO(base64.b64decode(response.data["zipBase64"]))) as archive:
        assert set(archive.namelist()) == {
            "contacts.vcf",
            "clean_contacts.xlsx",
            "merged_duplicates.xlsx",
            "invalid_rows.xlsx",
            "summary.txt",
        }
    assert AuditEvent.objects.filter(
        action="service.execute",
        actor=normal_user,
        target_id=str(excel_service.id),
        outcome="success",
    ).exists()


def test_invalid_extension_is_rejected(api, normal_user, excel_service):
    authenticate(api, normal_user)
    response = api.post(URL, {"file": upload(b"name,phone", "contacts.csv")}, format="multipart")
    assert response.status_code == 400
    assert response.data["code"] == "INVALID_FILE_EXTENSION"


def test_empty_upload_is_rejected(api, normal_user, excel_service):
    authenticate(api, normal_user)
    response = api.post(URL, {"file": upload(b"", "contacts.xlsx")}, format="multipart")
    assert response.status_code == 400
    assert response.data["code"] == "EMPTY_FILE"


def test_file_signature_must_match_extension(api, normal_user, excel_service):
    authenticate(api, normal_user)
    response = api.post(
        URL,
        {"file": upload(b"plain text pretending to be Excel", "contacts.xlsx")},
        format="multipart",
    )
    assert response.status_code == 400
    assert response.data["code"] == "INVALID_FILE_SIGNATURE"


def test_oversized_file_is_rejected(api, normal_user, excel_service):
    authenticate(api, normal_user)
    response = api.post(
        URL,
        {"file": upload(b"PK\x03\x04" + b"x" * MAX_FILE_SIZE, "contacts.xlsx")},
        format="multipart",
    )
    assert response.status_code == 413
    assert response.data["code"] == "FILE_TOO_LARGE"


def test_malformed_workbook_returns_safe_error(api, normal_user, excel_service):
    authenticate(api, normal_user)
    response = api.post(
        URL,
        {"file": upload(b"PK\x03\x04not-a-real-workbook", "contacts.xlsx")},
        format="multipart",
    )
    assert response.status_code == 400
    assert response.data["code"] == "INVALID_WORKBOOK"
    assert "not-a-real-workbook" not in response.data["message"]


def test_unexpected_processing_error_does_not_leak_details(api, normal_user, excel_service, monkeypatch):
    def fail(*args, **kwargs):
        raise RuntimeError(r"C:\private\contacts.xlsx secret parser failure")

    monkeypatch.setattr("reports.excel_contacts.application.process_workbook", fail)
    authenticate(api, normal_user)
    response = api.post(URL, {"file": upload()}, format="multipart")
    assert response.status_code == 500
    assert response.data["code"] == "EXCEL_CONTACTS_PROCESSING_FAILED"
    assert "private" not in response.data["message"]
    assert AuditEvent.objects.filter(action="service.execute", outcome="failure").exists()


def test_excessive_row_limit_is_enforced(api, normal_user, excel_service):
    authenticate(api, normal_user)
    response = api.post(
        URL,
        {"file": upload(make_xlsx(excessive_row=True))},
        format="multipart",
    )
    assert response.status_code == 400
    assert response.data["code"] == "INVALID_WORKBOOK"


def test_formula_like_values_are_escaped_in_export(api, normal_user, excel_service):
    authenticate(api, normal_user)
    data = make_xlsx([["@SUM(1,1)", "0999123456", "user@example.com"]])
    response = api.post(URL, {"file": upload(data)}, format="multipart")
    assert response.status_code == 200
    with zipfile.ZipFile(io.BytesIO(base64.b64decode(response.data["zipBase64"]))) as archive:
        workbook = load_workbook(io.BytesIO(archive.read("clean_contacts.xlsx")), data_only=False)
        sheet = workbook.active
        assert sheet["A2"].value == "'@SUM(1,1)"
        assert sheet["B2"].value.startswith("'+")
        workbook.close()
