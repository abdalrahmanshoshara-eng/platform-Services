"""Bounded in-memory Excel parsing and result-archive generation."""

import io
import re
import zipfile
from datetime import UTC, datetime

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .domain import (
    DUPLICATE_COLUMNS,
    EXPECTED_COLUMNS,
    INVALID_COLUMNS,
    VALID_COLUMNS,
    create_vcard,
    is_empty,
    normalize_header,
    process_contact_rows,
    safe_spreadsheet_value,
)

MAX_WORKSHEETS = 20
MAX_ROWS = 10_000
MAX_COLUMNS = 100
MAX_XLSX_ARCHIVE_ENTRIES = 2_000
MAX_XLSX_UNCOMPRESSED_SIZE = 100 * 1024 * 1024
MAX_OUTPUT_SIZE = 25 * 1024 * 1024


class WorkbookValidationError(Exception):
    """Expected, safe-to-report workbook rejection."""


def _validate_xlsx_container(data: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
                raise WorkbookValidationError("يحتوي ملف Excel على عدد مفرط من الأجزاء.")
            if sum(item.file_size for item in entries) > MAX_XLSX_UNCOMPRESSED_SIZE:
                raise WorkbookValidationError("حجم محتوى ملف Excel بعد فك الضغط يتجاوز الحد المسموح.")
            if any(item.flag_bits & 0x1 for item in entries):
                raise WorkbookValidationError("ملفات Excel المشفرة غير مدعومة.")
    except zipfile.BadZipFile as exc:
        raise WorkbookValidationError("ملف Excel تالف أو غير صالح.") from exc


def _non_empty_matrix(rows) -> list[list]:
    return [list(row) for row in rows if any(not is_empty(value) for value in row)]


def _read_xlsx(data: bytes) -> tuple[list[list], str]:
    _validate_xlsx_container(data)
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        if len(workbook.sheetnames) > MAX_WORKSHEETS:
            raise WorkbookValidationError(f"الحد الأعلى هو {MAX_WORKSHEETS} ورقة عمل.")
        if not workbook.sheetnames:
            raise WorkbookValidationError("ملف Excel لا يحتوي على أوراق عمل.")
        sheet = workbook[workbook.sheetnames[0]]
        if sheet.max_row > MAX_ROWS + 1:
            raise WorkbookValidationError(f"الحد الأعلى المدعوم هو {MAX_ROWS:,} صف لكل ملف.")
        if sheet.max_column > MAX_COLUMNS:
            raise WorkbookValidationError(f"الحد الأعلى المدعوم هو {MAX_COLUMNS} عمود لكل ورقة.")
        matrix = _non_empty_matrix(sheet.iter_rows(values_only=True))
        return matrix, sheet.title
    except WorkbookValidationError:
        raise
    except Exception as exc:
        raise WorkbookValidationError("ملف Excel تالف أو غير صالح.") from exc
    finally:
        if "workbook" in locals():
            workbook.close()


def _read_xls(data: bytes) -> tuple[list[list], str]:
    try:
        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
        if workbook.nsheets > MAX_WORKSHEETS:
            raise WorkbookValidationError(f"الحد الأعلى هو {MAX_WORKSHEETS} ورقة عمل.")
        if workbook.nsheets == 0:
            raise WorkbookValidationError("ملف Excel لا يحتوي على أوراق عمل.")
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows > MAX_ROWS + 1:
            raise WorkbookValidationError(f"الحد الأعلى المدعوم هو {MAX_ROWS:,} صف لكل ملف.")
        if sheet.ncols > MAX_COLUMNS:
            raise WorkbookValidationError(f"الحد الأعلى المدعوم هو {MAX_COLUMNS} عمود لكل ورقة.")
        matrix = _non_empty_matrix(sheet.row_values(index) for index in range(sheet.nrows))
        return matrix, sheet.name
    except WorkbookValidationError:
        raise
    except Exception as exc:
        raise WorkbookValidationError("ملف Excel تالف أو غير صالح.") from exc
    finally:
        if "workbook" in locals():
            workbook.release_resources()


def _extract_rows(matrix: list[list]) -> list[dict]:
    if not matrix:
        raise WorkbookValidationError("ورقة العمل الأولى فارغة.")
    headers = [normalize_header(value) for value in matrix[0]]
    missing = [column for column in EXPECTED_COLUMNS if column not in headers]
    if missing:
        raise WorkbookValidationError(f"الأعمدة المطلوبة غير موجودة: {'، '.join(missing)}.")
    index_by_header = {header: index for index, header in enumerate(headers)}
    return [
        {
            column: cells[index_by_header[column]] if index_by_header[column] < len(cells) else ""
            for column in EXPECTED_COLUMNS
        }
        for cells in matrix[1:]
    ]


def _create_excel_report(sheet_name: str, columns: list[str], rows: list[dict], header_color: str) -> bytes:
    workbook = Workbook()
    workbook.creator = "Professional Reports Platform"
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.sheet_view.rightToLeft = True
    sheet.freeze_panes = "A2"
    sheet.append(columns)
    for row in rows:
        sheet.append([safe_spreadsheet_value(row.get(column, "")) for column in columns])
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(1, len(rows) + 1)}"

    fill = PatternFill(fill_type="solid", fgColor=header_color)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="1F2937")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_index, column in enumerate(columns, start=1):
        longest = max([len(str(row.get(column, ""))) for row in rows] + [len(column)])
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(longest + 3, 14), 48)

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _safe_source_filename(filename: str) -> str:
    basename = str(filename).replace("\\", "/").split("/")[-1]
    return re.sub(r"[\x00-\x1f\x7f]", "", basename)[:200] or "contacts.xlsx"


def _summary_text(source_filename: str, sheet_name: str, country_code: str, summary: dict) -> str:
    created_at = datetime.now(UTC).isoformat(timespec="seconds")
    return "\n".join(
        [
            "تقرير تحويل جهات الاتصال",
            f"تاريخ الإنشاء (UTC): {created_at}",
            f"الملف المصدر: {_safe_source_filename(source_filename)}",
            f"ورقة العمل المستخدمة: {sheet_name}",
            f"إجمالي الصفوف: {summary['totalRows']}",
            f"جهات الاتصال الفريدة المقبولة: {summary['validCount']}",
            f"السجلات المكررة التي تم دمجها: {summary['duplicateCount']}",
            f"الصفوف التي تحتاج مراجعة: {summary['invalidCount']}",
            "طريقة إزالة التكرار: الاحتفاظ بأول ظهور لكل رقم تواصل بعد توحيد صيغته",
            f"رمز الدولة الافتراضي للأرقام المحلية: +{country_code}",
            "",
        ]
    )


def process_workbook(data: bytes, extension: str, source_filename: str, country_code: str) -> dict:
    matrix, sheet_name = _read_xlsx(data) if extension == ".xlsx" else _read_xls(data)
    rows = _extract_rows(matrix)
    if len(rows) > MAX_ROWS:
        raise WorkbookValidationError(f"الحد الأعلى المدعوم هو {MAX_ROWS:,} صف لكل ملف.")
    result = process_contact_rows(rows, country_code)

    archive_buffer = io.BytesIO()
    with zipfile.ZipFile(archive_buffer, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as archive:
        archive.writestr("contacts.vcf", create_vcard(result["valid_rows"]))
        archive.writestr(
            "clean_contacts.xlsx",
            _create_excel_report("جهات الاتصال الفريدة", VALID_COLUMNS, result["valid_rows"], "D9EAD3"),
        )
        archive.writestr(
            "merged_duplicates.xlsx",
            _create_excel_report("مكررات تم دمجها", DUPLICATE_COLUMNS, result["duplicate_rows"], "FCE5CD"),
        )
        archive.writestr(
            "invalid_rows.xlsx",
            _create_excel_report("صفوف تحتاج مراجعة", INVALID_COLUMNS, result["invalid_rows"], "F4CCCC"),
        )
        archive.writestr(
            "summary.txt",
            _summary_text(source_filename, sheet_name, country_code, result["summary"]),
        )

    zip_buffer = archive_buffer.getvalue()
    if len(zip_buffer) > MAX_OUTPUT_SIZE:
        raise WorkbookValidationError("حجم ملف النتائج يتجاوز الحد المسموح.")
    return {
        **result,
        "zip_buffer": zip_buffer,
        "source_sheet_name": sheet_name,
    }
